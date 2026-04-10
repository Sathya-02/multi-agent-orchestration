"""
tools.py — Agent tools
Includes: MockSearchTool, DataAnalysisTool, SummaryTool,
          FileReadTool (reads uploaded files), MathTool (evaluates expressions),
          SpawnAgentTool (agent requests a new agent — human approval required,
                          disabled when spawn toggle is off, blocked on duplicate roles).
"""
from langchain.tools import BaseTool
from typing import Any
import sys
import time, random, json, math, re
from pathlib import Path

UPLOADS_DIR = Path(__file__).parent / "uploads"
UPLOADS_DIR.mkdir(exist_ok=True)


def to_str(data: Any) -> str:
    if isinstance(data, dict):
        for key in ("query", "data", "text", "input", "content", "expression", "filename"):
            if key in data:
                return str(data[key])
        return json.dumps(data)
    return str(data) if data else ""


def _sync_broadcast(msg: dict) -> None:
    """
    Safe late-binding wrapper for main.sync_broadcast.

    Uses sys.modules instead of 'import main' to avoid the circular-import /
    partial-initialisation race where main is not yet fully loaded when tools.py
    is first imported.  By the time any tool's _run() is actually called the
    main module is always fully initialised, so sys.modules['main'] is safe.
    """
    _main = sys.modules.get("main")
    if _main is None:
        return  # main not loaded yet — drop the message silently
    fn = getattr(_main, "sync_broadcast", None)
    if callable(fn):
        fn(msg)


# ── Web search — real or mock depending on config ────────────────────────
class MockSearchTool(BaseTool):
    """
    Smart web search tool — uses real-time providers when enabled,
    falls back to mock results when not configured.

    Real providers (no API key needed):
      • DuckDuckGo  — general search, news         (pip install duckduckgo-search)
      • Wikipedia   — factual / encyclopaedic       (built-in REST)
      • wttr.in     — weather queries               (built-in REST)
      • WorldTimeAPI— current date / time / TZ      (built-in REST)
      • ExchangeRate— currency / FX rates           (built-in REST)

    Enable via ⚙️ Settings → 🌐 Web Search in the UI, or:
      POST /web-search/config  {"enabled": true}
    """
    name: str        = "web_search"
    description: str = (
        "Search the web for real-time and current information. "
        "ALWAYS call this tool for: today's date, current day of the week, "
        "current time, weather in any location, latest news, current events, "
        "live exchange rates, stock prices, sports scores, and anything that changes. "
        "NEVER guess, estimate, or write placeholder text for real-time data. "
        "Input: a plain-English query, e.g. 'what day is today', "
        "'weather in Chennai', 'USD to INR rate', 'latest AI news'."
    )

    def _run(self, query: Any = None, **kwargs) -> str:
        q = to_str(query or kwargs) or "general research"
        try:
            from web_search_tool import real_search, load_config as _load_ws_cfg
            cfg = _load_ws_cfg()
            if cfg.get("enabled", False):
                return real_search(q)
        except Exception as e:
            # If real search module fails for any reason, fall through to mock
            pass
        # Mock fallback
        time.sleep(0.3)
        return (
            f"Search results for \'{q}\' [mock — enable real search in Settings]:\n"
            f"• Finding 1: Recent studies show significant trends related to {q}.\n"
            f"• Finding 2: Expert analysis identifies multiple perspectives on {q}.\n"
            f"• Finding 3: Statistical overview indicates growing interest in {q}."
        )

    def _arun(self, *a, **kw): raise NotImplementedError


# ── Data analysis ─────────────────────────────────────────────────────────
class DataAnalysisTool(BaseTool):
    name: str        = "data_analyser"
    description: str = "Analyse data or text and extract key insights. Input: text or data to analyse."

    def _run(self, data: Any = None, **kwargs) -> str:
        text = to_str(data or kwargs)
        time.sleep(0.3)
        return (
            f"Analysis complete.\n"
            f"• Key theme: {text[:60]}…\n"
            f"• Sentiment: Neutral-Positive\n"
            f"• Confidence: {random.randint(72, 96)}%\n"
            f"• Recommended action: Proceed with synthesis"
        )

    def _arun(self, *a, **kw): raise NotImplementedError


# ── Summariser ────────────────────────────────────────────────────────────
class SummaryTool(BaseTool):
    name: str        = "summariser"
    description: str = "Summarise long content into concise bullet points. Input: text to summarise."

    def _run(self, data: Any = None, **kwargs) -> str:
        text  = to_str(data or kwargs)
        words = text.split()[:30]
        time.sleep(0.2)
        return f"Summary: {' '.join(words)}… [condensed to key points]"

    def _arun(self, *a, **kw): raise NotImplementedError


# ── File reader ───────────────────────────────────────────────────────────
class FileReadTool(BaseTool):
    name: str        = "read_uploaded_file"
    description: str = (
        "Read the content of an uploaded file. "
        "Input: filename (e.g. 'report.pdf', 'data.csv', 'notes.txt'). "
        "Returns the full text content of the file."
    )

    def _run(self, filename: Any = None, **kwargs) -> str:
        fname = to_str(filename or kwargs).strip()
        if not fname:
            files = list(UPLOADS_DIR.glob("*"))
            if not files:
                return "No files have been uploaded yet."
            return "Available uploaded files:\n" + "\n".join(f.name for f in files)

        safe  = re.sub(r'[^a-zA-Z0-9._\- ]', '', fname).strip()
        path  = UPLOADS_DIR / safe
        if not path.exists():
            matches = [f for f in UPLOADS_DIR.glob("*") if f.name.lower() == safe.lower()]
            if not matches:
                avail = [f.name for f in UPLOADS_DIR.glob("*")]
                return f"File '{safe}' not found. Available: {avail}"
            path = matches[0]

        try:
            ext = path.suffix.lower()

            if ext in (".txt", ".md", ".csv", ".json", ".log", ".yaml", ".yml"):
                content = path.read_text(encoding="utf-8", errors="replace")
                if len(content) > 8000:
                    content = content[:8000] + "\n\n[... truncated — showing first 8000 chars]"
                return f"=== Content of {path.name} ===\n\n{content}"

            if ext == ".pdf":
                try:
                    import pypdf
                    reader = pypdf.PdfReader(str(path))
                    text   = "\n".join(p.extract_text() or "" for p in reader.pages)
                    if len(text) > 8000:
                        text = text[:8000] + "\n[truncated]"
                    return f"=== PDF: {path.name} ===\n\n{text}"
                except ImportError:
                    return f"PDF uploaded ({path.name}) but pypdf not installed. Run: pip install pypdf"

            if ext == ".docx":
                try:
                    import docx
                    doc  = docx.Document(str(path))
                    text = "\n".join(p.text for p in doc.paragraphs)
                    if len(text) > 8000:
                        text = text[:8000] + "\n[truncated]"
                    return f"=== DOCX: {path.name} ===\n\n{text}"
                except ImportError:
                    return f"DOCX uploaded ({path.name}) but python-docx not installed. Run: pip install python-docx"

            if ext in (".xlsx", ".xls"):
                try:
                    import openpyxl
                    wb  = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
                    out = []
                    for sheet in wb.sheetnames:
                        ws   = wb[sheet]
                        rows = list(ws.iter_rows(values_only=True))[:50]
                        out.append(f"Sheet: {sheet}")
                        out.extend(", ".join(str(c) for c in r) for r in rows)
                    return f"=== Excel: {path.name} ===\n\n" + "\n".join(out)
                except ImportError:
                    return f"Excel uploaded ({path.name}) but openpyxl not installed. Run: pip install openpyxl"

            if ext in (".png", ".jpg", ".jpeg", ".gif", ".webp"):
                size = path.stat().st_size
                return f"Image file: {path.name} ({size} bytes). Image analysis not supported."

            try:
                content = path.read_text(encoding="utf-8", errors="replace")[:4000]
                return f"=== {path.name} ===\n\n{content}"
            except Exception:
                return f"File '{path.name}' could not be read as text (binary format)."

        except Exception as e:
            return f"Error reading '{fname}': {e}"

    def _arun(self, *a, **kw): raise NotImplementedError


# ── Math evaluator ────────────────────────────────────────────────────────
class MathTool(BaseTool):
    name: str        = "calculator"
    description: str = (
        "Evaluate mathematical expressions and perform calculations. "
        "Input: a mathematical expression as a string, e.g. '2 + 2', 'sqrt(144)', '15% of 3200'."
    )

    def _run(self, expression: Any = None, **kwargs) -> str:
        expr = to_str(expression or kwargs).strip()
        if not expr:
            return "No expression provided."

        pct = re.match(r"([\d.]+)\s*%\s*of\s*([\d.]+)", expr, re.IGNORECASE)
        if pct:
            a, b = float(pct.group(1)), float(pct.group(2))
            return f"{a}% of {b} = {a * b / 100}"

        safe_ns = {k: getattr(math, k) for k in dir(math) if not k.startswith("_")}
        safe_ns["abs"] = abs
        safe_ns["round"] = round
        try:
            if re.search(r"[^0-9\s\+\-\*/\.\(\)\^%sqrtlogpiecosintanabsround,]", expr.lower()):
                return f"Expression contains unsupported characters: {expr}"
            result = eval(expr.replace("^", "**"), {"__builtins__": {}}, safe_ns)  # noqa: S307
            return f"{expr} = {result}"
        except Exception as e:
            return f"Could not evaluate '{expr}': {e}"

    def _arun(self, *a, **kw): raise NotImplementedError


# ── Spawn agent tool ──────────────────────────────────────────────────────
class SpawnAgentTool(BaseTool):
    name: str        = "request_new_agent"
    description: str = (
        "Request the creation of a new specialised agent when the current team "
        "lacks a needed capability. Input: a JSON string with keys: "
        "'role', 'goal', 'backstory', 'reason'. "
        "The request will be sent for human approval before the agent is created."
    )

    def _run(self, suggestion: Any = None, **kwargs) -> str:
        from agent_registry import (
            is_spawn_enabled, role_exists, find_agent_by_role, request_spawn,
        )

        # ── Check 1: spawn toggle ─────────────────────────────────────────
        if not is_spawn_enabled():
            return (
                "Agent spawning is currently disabled by the operator. "
                "Use the existing team members to complete this task."
            )

        # ── Parse suggestion ──────────────────────────────────────────────
        raw = to_str(suggestion or kwargs)
        try:
            data = json.loads(raw)
        except Exception:
            data = {"role": raw, "goal": "Complete specialised tasks.",
                    "backstory": raw, "reason": raw}

        requested_role = data.get("role", "").strip()

        # ── Check 2: role deduplication ───────────────────────────────────
        if requested_role and role_exists(requested_role):
            existing = find_agent_by_role(requested_role)
            return (
                f"An agent with the role '{requested_role}' already exists "
                f"({existing['icon']} {existing['label']}, id: {existing['id']}). "
                f"Please delegate this task to that agent instead of requesting a duplicate."
            )

        # ── Submit spawn request ──────────────────────────────────────────
        req = request_spawn(requested_by="agent", suggestion=data)

        _sync_broadcast({
            "type":       "spawn_request",
            "request_id": req["request_id"],
            "suggestion": data,
            "message": (
                f"🤖 Agent requests new specialist: '{requested_role or '?'}' — "
                f"awaiting human approval"
            ),
        })

        return (
            f"Spawn request submitted (ID: {req['request_id']}) for "
            f"'{requested_role or 'Unknown'}'. Awaiting human approval."
        )

    def _arun(self, *a, **kw): raise NotImplementedError



# ── Spawn tool tool ───────────────────────────────────────────────────────
class SpawnToolTool(BaseTool):
    name: str        = "request_new_tool"
    description: str = (
        "Request the creation of a new custom tool when a capability you need "
        "is missing from your available tools. "
        "Input: a JSON string with keys: 'name' (snake_case tool id), "
        "'display_name', 'description', 'code' (Python body of _run), 'reason'. "
        "The request will be sent for human approval before the tool is created."
    )

    def _run(self, suggestion: Any = None, **kwargs) -> str:
        from tool_registry import (
            name_exists, find_tool_by_name, request_tool_spawn,
        )

        raw = to_str(suggestion or kwargs)
        try:
            data = json.loads(raw)
        except Exception:
            data = {"name": _slug_name(raw), "description": raw,
                    "display_name": raw, "code": "    return str(input_data)", "reason": raw}

        req_name = data.get("name", "").strip()

        # Deduplicate
        if req_name and name_exists(req_name):
            existing = find_tool_by_name(req_name)
            return (
                f"A tool named '{req_name}' already exists "
                f"({existing.get('display_name','?')}). "
                "Use that tool instead of requesting a duplicate."
            )

        req = request_tool_spawn(requested_by="agent", suggestion=data)

        _sync_broadcast({
            "type":          "tool_spawn_request",
            "request_id":    req["request_id"],
            "suggestion":    data,
            "message": (
                f"🔧 Agent requests new tool: '{req_name or '?'}' — "
                "awaiting human approval"
            ),
        })

        return (
            f"Tool spawn request submitted (ID: {req['request_id']}) for "
            f"'{req_name or 'unknown'}'. Awaiting human approval."
        )

    def _arun(self, *a, **kw): raise NotImplementedError


def _slug_name(text: str) -> str:
    import re as _re
    s = _re.sub(r'[^a-zA-Z0-9 _-]', '', text.strip().lower())
    return _re.sub(r'[ \-]+', '_', s).strip('_') or "custom_tool"


# ── Knowledge Base / RAG search tool ─────────────────────────────────────
class KnowledgeBaseSearchTool(BaseTool):
    name: str        = "knowledge_base_search"
    description: str = (
        "Search the local knowledge base for relevant information from ingested documents. "
        "Use this tool BEFORE web_search when answering questions about topics that may be "
        "covered in uploaded documents, company docs, or previously ingested knowledge. "
        "Input: a natural-language query. "
        "Returns: relevant text chunks with source file and relevance score."
    )

    def _run(self, query: Any = None, **kwargs) -> str:
        q = to_str(query or kwargs) or "general"
        try:
            from rag_engine import search
            return search(q)
        except Exception as e:
            return f"Knowledge base search error: {e}"

    def _arun(self, *a, **kw): raise NotImplementedError

# ── Tool factory ──────────────────────────────────────────────────────────
def make_tools(tool_names: list[str]) -> list:
    """Instantiate tools by name — includes builtin + custom tools from registry."""
    from tool_registry import instantiate_tool, get_tool as _get_tool
    builtin_registry = {
        "web_search":         MockSearchTool,
        "data_analyser":      DataAnalysisTool,
        "summariser":         SummaryTool,
        "read_uploaded_file": FileReadTool,
        "calculator":         MathTool,
        "request_new_agent":  SpawnAgentTool,
        "request_new_tool":          SpawnToolTool,
        "knowledge_base_search":     KnowledgeBaseSearchTool,
    }
    tools = []
    for n in tool_names:
        if n in builtin_registry:
            tools.append(builtin_registry[n]())
        else:
            # Try to load as a custom tool from registry
            inst = instantiate_tool(n)
            if inst is not None:
                tools.append(inst)
    return tools
