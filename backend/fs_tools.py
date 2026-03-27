"""
fs_tools.py — Filesystem agent tools
Provides four tools for the File System Agent:

  fs_read_file   — read any text/binary file from an allowed folder
  fs_write_file  — create a new file in an allowed folder (write permission required)
  fs_edit_file   — overwrite or append to an existing file (edit permission required)
  fs_list_dir    — list files and subdirectories in an allowed folder (read permission)

All tools enforce the access control list in fs_config.py.
By default no write or edit access is granted anywhere.
"""
from langchain.tools import BaseTool
from typing import Any
from pathlib import Path
import json, re

from fs_config import can_read, can_write, can_edit, _resolve_safe, get_accessible_paths

TEXT_EXTS = {
    ".txt", ".md", ".csv", ".json", ".jsonl", ".yaml", ".yml",
    ".log", ".py", ".js", ".jsx", ".ts", ".tsx", ".html", ".css",
    ".sh", ".bash", ".toml", ".ini", ".cfg", ".conf", ".xml",
    ".sql", ".r", ".rb", ".go", ".rs", ".java", ".c", ".cpp",
    ".h", ".env", ".gitignore",
}
MAX_READ_CHARS = 12_000


def _to_str(data: Any) -> str:
    if isinstance(data, dict):
        for k in ("path", "filename", "file", "content", "text", "input"):
            if k in data:
                return str(data[k])
        return json.dumps(data)
    return str(data) if data else ""


def _parse_input(raw: Any) -> dict:
    """Try to parse JSON input; fall back to {path: raw_string}."""
    s = _to_str(raw)
    try:
        d = json.loads(s)
        return d if isinstance(d, dict) else {"path": s}
    except Exception:
        return {"path": s}


# ── FSReadTool ─────────────────────────────────────────────────────────────
class FSReadTool(BaseTool):
    name: str        = "fs_read_file"
    description: str = (
        "Read the contents of a file from the local filesystem. "
        "Input: either a JSON string {\"path\": \"/absolute/path/to/file.txt\"} "
        "OR just the plain file path as a string. "
        "Call with no input to list all accessible folders and their permissions."
    )

    def _run(self, path: Any = None, **kwargs) -> str:
        # No input → list what's accessible
        if path is None or (isinstance(path, str) and not path.strip()):
            accessible = get_accessible_paths()
            if not accessible:
                return (
                    "No filesystem folders are configured for agent access.\n"
                    "Ask the operator to add folders via the 📁 Filesystem panel in the UI."
                )
            return (
                "Accessible folders (add these as path prefixes when reading files):\n"
                + "\n".join(f"  • {p}" for p in accessible)
            )

        inp   = _parse_input(path or kwargs)
        fpath = inp.get("path", "").strip()

        if not fpath:
            accessible = get_accessible_paths()
            return (
                "No path provided.\n"
                f"Accessible folders: {accessible}\n"
                "Usage: {\"path\": \"/full/path/to/file.txt\"}"
            )

        ok, msg = can_read(fpath, agent="fs_agent")
        if not ok:
            return f"❌ Permission denied: {msg}"

        p = _resolve_safe(fpath)
        if p is None:
            return "❌ Invalid path."
        if not p.exists():
            # Try without resolve (for macOS symlink edge case)
            from pathlib import Path as _Path
            p2 = _Path(fpath).expanduser()
            if p2.exists():
                p = p2
            else:
                return f"❌ File not found: {fpath}"
        if p.is_dir():
            return f"❌ '{p}' is a directory. Use fs_list_dir to list its contents."

        try:
            ext = p.suffix.lower()

            # PDF
            if ext == ".pdf":
                try:
                    import pypdf
                    reader = pypdf.PdfReader(str(p))
                    text   = "\n".join(page.extract_text() or "" for page in reader.pages)
                    if len(text) > MAX_READ_CHARS:
                        text = text[:MAX_READ_CHARS] + "\n[... truncated]"
                    return f"=== PDF: {p.name} ===\n\n{text}"
                except ImportError:
                    return "pypdf not installed. Run: pip install pypdf"

            # DOCX
            if ext == ".docx":
                try:
                    import docx
                    doc  = docx.Document(str(p))
                    text = "\n".join(para.text for para in doc.paragraphs)
                    if len(text) > MAX_READ_CHARS:
                        text = text[:MAX_READ_CHARS] + "\n[... truncated]"
                    return f"=== DOCX: {p.name} ===\n\n{text}"
                except ImportError:
                    return "python-docx not installed. Run: pip install python-docx"

            # Excel
            if ext in (".xlsx", ".xls"):
                try:
                    import openpyxl
                    wb  = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
                    out = []
                    for sheet in wb.sheetnames:
                        ws   = wb[sheet]
                        rows = list(ws.iter_rows(values_only=True))[:100]
                        out.append(f"\nSheet: {sheet}")
                        out.extend(", ".join(str(c) for c in r) for r in rows)
                    return f"=== Excel: {p.name} ===\n" + "\n".join(out)
                except ImportError:
                    return "openpyxl not installed. Run: pip install openpyxl"

            # Text / code
            if ext in TEXT_EXTS or ext == "":
                content = p.read_text(encoding="utf-8", errors="replace")
                if len(content) > MAX_READ_CHARS:
                    content = content[:MAX_READ_CHARS] + "\n\n[... truncated — file too large]"
                return f"=== {p} ===\n\n{content}"

            # Unknown — try text, fall back to binary info
            try:
                content = p.read_text(encoding="utf-8", errors="replace")[:MAX_READ_CHARS]
                return f"=== {p} ===\n\n{content}"
            except Exception:
                size = p.stat().st_size
                return f"Binary file: {p.name} ({size:,} bytes) — cannot read as text."

        except Exception as e:
            return f"❌ Error reading '{p}': {e}"

    def _arun(self, *a, **kw): raise NotImplementedError


# ── FSListTool ─────────────────────────────────────────────────────────────
class FSListTool(BaseTool):
    name: str        = "fs_list_dir"
    description: str = (
        "List files and subdirectories in a folder on the local filesystem. "
        "Input: JSON {\"path\": \"/folder\", \"pattern\": \"*.txt\"} (pattern optional), "
        "OR just the folder path as a plain string. "
        "Call with no input to see all accessible folders."
    )

    def _run(self, path: Any = None, **kwargs) -> str:
        # No input → show accessible folders
        if path is None or (isinstance(path, str) and not path.strip()):
            accessible = get_accessible_paths()
            if not accessible:
                return (
                    "No folders are configured for agent access.\n"
                    "Ask the operator to add a folder in the 📁 Filesystem panel."
                )
            return (
                "Accessible folders:\n"
                + "\n".join(f"  • {p}" for p in accessible)
            )

        inp     = _parse_input(path or kwargs)
        dpath   = inp.get("path", "").strip()
        pattern = inp.get("pattern", "*").strip() or "*"

        if not dpath:
            return "No path provided."

        ok, msg = can_read(dpath, agent="fs_agent")
        if not ok:
            return f"❌ Permission denied: {msg}"

        p = _resolve_safe(dpath)
        if p is None:
            return "❌ Invalid path."

        # Try both expanded and resolved forms if first doesn't exist
        if not p.exists():
            from pathlib import Path as _Path
            p2 = _Path(dpath).expanduser()
            if p2.exists():
                p = p2
            else:
                return f"❌ Directory not found: {dpath}"
        if not p.is_dir():
            return f"❌ '{p}' is a file, not a directory. Use fs_read_file to read it."

        try:
            items = sorted(p.glob(pattern))
            if not items:
                return f"Directory '{p}' is empty (pattern: {pattern})."

            lines = [f"📁 {p}  ({len(items)} items)\n"]
            for item in items[:200]:
                rel   = item.relative_to(p)
                icon  = "📁" if item.is_dir() else "📄"
                size  = f"  {item.stat().st_size:>10,} B" if item.is_file() else ""
                lines.append(f"  {icon} {rel}{size}")
            if len(items) > 200:
                lines.append(f"  … and {len(items) - 200} more items")
            return "\n".join(lines)
        except Exception as e:
            return f"❌ Error listing '{p}': {e}"

    def _arun(self, *a, **kw): raise NotImplementedError


# ── FSWriteTool ────────────────────────────────────────────────────────────
class FSWriteTool(BaseTool):
    name: str        = "fs_write_file"
    description: str = (
        "Create a NEW file on the local filesystem. "
        "Will NOT overwrite an existing file — use fs_edit_file for that. "
        "Input JSON: {\"path\": \"/absolute/path/to/new_file.txt\", \"content\": \"file content here\"}. "
        "Write access must be granted for the parent folder."
    )

    def _run(self, path: Any = None, **kwargs) -> str:
        inp     = _parse_input(path or kwargs)
        fpath   = inp.get("path", "").strip()
        content = inp.get("content", inp.get("text", inp.get("data", "")))

        if not fpath:
            return "No path provided."
        if content is None:
            return "No content provided. Input: {\"path\": \"...\", \"content\": \"...\"}"

        ok, msg = can_write(fpath, agent="fs_agent")
        if not ok:
            return f"❌ Permission denied: {msg}"

        p = _resolve_safe(fpath)
        if p is None:
            return "❌ Invalid path."
        # Use expanduser form for existence check (avoids macOS symlink issue)
        from pathlib import Path as _Path
        p_exp = _Path(fpath).expanduser()
        if p_exp.exists() or p.exists():
            return (
                f"❌ File already exists: {fpath}\n"
                "Use fs_edit_file to overwrite or append to an existing file."
            )

        try:
            p.parent.mkdir(parents=True, exist_ok=True)
            p.write_text(str(content), encoding="utf-8")
            return f"✅ File created: {p}  ({p.stat().st_size:,} bytes)"
        except Exception as e:
            return f"❌ Error writing '{p}': {e}"

    def _arun(self, *a, **kw): raise NotImplementedError


# ── FSEditTool ─────────────────────────────────────────────────────────────
class FSEditTool(BaseTool):
    name: str        = "fs_edit_file"
    description: str = (
        "Edit an EXISTING file on the local filesystem — overwrite it completely or append to it. "
        "Input JSON: {"
        "\"path\": \"/absolute/path/to/file.txt\", "
        "\"content\": \"new content\", "
        "\"mode\": \"overwrite\"  (or \"append\")"
        "}. "
        "Edit access must be granted for the folder. "
        "Default mode is 'overwrite'."
    )

    def _run(self, path: Any = None, **kwargs) -> str:
        inp     = _parse_input(path or kwargs)
        fpath   = inp.get("path", "").strip()
        content = inp.get("content", inp.get("text", inp.get("data", "")))
        mode    = inp.get("mode", "overwrite").strip().lower()

        if not fpath:
            return "No path provided."
        if content is None:
            return "No content provided."
        if mode not in ("overwrite", "append"):
            mode = "overwrite"

        ok, msg = can_edit(fpath, agent="fs_agent")
        if not ok:
            return f"❌ Permission denied: {msg}"

        p = _resolve_safe(fpath)
        if p is None:
            return "❌ Invalid path."
        if not p.exists():
            return (
                f"❌ File not found: {p}\n"
                "Use fs_write_file to create a new file."
            )
        if not p.is_file():
            return f"❌ '{p}' is not a file."

        try:
            original_size = p.stat().st_size
            if mode == "append":
                with p.open("a", encoding="utf-8") as f:
                    f.write(str(content))
                new_size = p.stat().st_size
                added    = new_size - original_size
                return f"✅ Appended {added:,} bytes to {p}  (total: {new_size:,} bytes)"
            else:
                p.write_text(str(content), encoding="utf-8")
                new_size = p.stat().st_size
                return f"✅ Overwrote {p}  ({new_size:,} bytes, was {original_size:,} bytes)"
        except Exception as e:
            return f"❌ Error editing '{p}': {e}"

    def _arun(self, *a, **kw): raise NotImplementedError
